import os
import numpy as np
import pandas as pd
from flask import Flask, request, render_template, redirect, url_for, send_file, session, flash
from werkzeug.utils import secure_filename
import pickle
from openpyxl import load_workbook
from ML_scripts.knn import knn
from ML_scripts.logistic import logistic
from ML_scripts.random_forest import random_forest
from ML_scripts.SVM import SVM
from ML_scripts.tensor_nn import tensorflow_nn

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['RESULTS_FOLDER'] = 'results'
app.config['NEW_MODELS'] = 'new_models'
ALLOWED_EXTENSIONS = {"csv"}
app.config['SECRET_KEY'] = 'my5EWRT$^W464'

## Data Pre Precessing Codes
def multipleLines(x):
    if x == 'Yes':
        return 1
    else:
        return 0

def is_float(x):
    try:
        float(x)
        return True
    except ValueError:
        return False
    
def convert(dic, x):
    if x not in dic:
        raise Exception('Dictionary has error, try to re-run the whole script')
    else:
        return dic[x]
    
def tenure_edit(x):
    if 0 <= x <= 9:
        return 0
    elif 29 >= x > 9:
        return 1
    elif 55 >= x > 29:
        return 2
    elif x > 55:
        return 3
    else:
        return 'problem'

def churn(x):
    if x == 1:
        return "Churn"
    else:
        return "Does Not Churn"

def ones_and_zeros(x):
    if x.lower() == 'yes':
        return 1
    elif x.lower() == 'no':
        return 0

# defining allowed files
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# preprocessing fixed code
def process_files(file_path):
    df = pd.read_csv(file_path)
    #one hot encoding for categorical variables
    df = df.drop(columns = ['customerID'])
    df = pd.concat([df, pd.get_dummies(df['PaymentMethod'])], axis = 1)
    df = df.drop(columns = ['PaymentMethod'])
    df = pd.concat([df, pd.get_dummies(df['Contract'])], axis = 1)
    df = df.drop(columns = ['Contract'])

    #converting to binary variables
    df['MultipleLines'] = df['MultipleLines'].apply(lambda x: multipleLines(x))
    df['OnlineSecurity'] = df['OnlineSecurity'].apply(lambda x: multipleLines(x))
    df['OnlineBackup'] = df['OnlineBackup'].apply(lambda x: multipleLines(x))
    df['DeviceProtection'] = df['DeviceProtection'].apply(lambda x: multipleLines(x))
    df['TechSupport'] = df['TechSupport'].apply(lambda x: multipleLines(x))
    df['StreamingTV'] = df['StreamingTV'].apply(lambda x: multipleLines(x))
    df['StreamingMovies'] = df['StreamingMovies'].apply(lambda x: multipleLines(x))
    df = pd.concat([df, pd.get_dummies(df['InternetService'])], axis = 1)
    df = df.rename(columns = {'No': 'No Internet Service'})
    df = df.drop(columns = ['InternetService', 'Bank transfer (automatic)', 'Month-to-month'])

    #data conversion from str to float/integer
    col_names = {}
    for i in df:
        col_names[i] = {}
        for j in df[i]:
            if isinstance(j, str) and not is_float(j):
                if j not in col_names[i]:
                    col_names[i][j] = len(col_names[i])
            else:
                if isinstance(j, int):
                    col_names[i][j] = j
                else:
                    col_names[i][j] = float(j)
    for i in df:
        df[i] = df[i].apply(lambda x: convert(col_names[i], x))

    #converting "tenure" into ordinal variable
    df['tenure'] = df['tenure'].apply(tenure_edit)
    return df

@app.route('/')
def home():
    return render_template('start.html')

@app.route('/index_two')
def index_two():
    return render_template('submit_training.html')

@app.route('/submit_train', methods = ['GET', 'POST'])
def submit_train():
    if 'file' not in request.files:
        return redirect(url_for('index_two'))

    file = request.files['file']
    if file.filename == '':
        return redirect(url_for('index_two'))

    if file and allowed_file(file.filename):
        df = pd.read_csv(file)
        x = df.drop(columns = ['Churn'])
        x.to_csv('uploads/training.csv', index = False)
        x = process_files('uploads/training.csv')
        y = df['Churn']
        y = y.apply(ones_and_zeros)
        x.to_csv('uploads/training.csv', index = False)
        y.to_csv('uploads/test.csv', index = False)
        return render_template('train_model.html')
    else:
        flash("File Error! Submit Only CSV file formats")
        return redirect(url_for('index_two'))

@app.route('/train_models', methods = ['GET', 'POST'])
def train_models():
    model_performance = pd.DataFrame(columns=['Model', 'Accuracy', 'Precision', 'Recall', 'F1-Score'])
    x = 'uploads/training.csv'
    y = 'uploads/test.csv'
    df = knn(x,y,model_performance)
    print("##----------knn done----------(1/5 models trained)")
    df = logistic(x,y,df)
    print("##----------logistic regression done----------(2/5 models trained)")
    df = random_forest(x,y,df)
    print("##----------random forest done----------(3/5 models trained)")
    df = SVM(x,y,df)
    print("##----------svm done----------(4/5 models trained)")
    df = tensorflow_nn(x,y,df)
    print("##-----------TF done----------(5/5 models trained)")
    df = df.style.background_gradient(cmap='RdYlBu_r').format({
        'Accuracy': '{:.2%}',
        'Precision': '{:.2%}',
        'Recall': '{:.2%}',
        'F1-Score': '{:.2%}'
    })
    df.to_excel('results/model_evaluation.xlsx', index = False)
    return render_template('evaluate_model.html')

@app.route('/evaluate_models', methods = ['GET', 'POST'])
def evaluate_models():
    filepath = 'results/model_evaluation.xlsx'
    book = load_workbook(filepath)
    sheet = book.active
    if request.method == 'GET':
        return render_template('view_evaluation.html', data = sheet)
    else:
        return redirect(url_for('evaluate_models'))

@app.route('/download_evaluation', methods = ['GET', 'POST'])
def download_evaluation():
    results_filepath = 'results/model_evaluation.xlsx'
    if request.method == 'GET':
        if not os.path.exists(results_filepath):
            return redirect(url_for('evaluate_models'))
        else:
            return send_file(results_filepath, as_attachment=True)

@app.route('/new_models', methods=['GET', 'POST'])
def new_models():
    if len(os.listdir(app.config['NEW_MODELS'])) == 0:
        flash("No NEW models trained")
        return render_template('index.html')
    if request.method == 'POST':
        filepath = session.get('filepath')
        if "csv" not in filepath:
            return redirect(url_for('index'))

        df = process_files(filepath)

        # Get the selected model option
        selected_model = request.form.get('predict_model')

        if selected_model is None:
            return redirect(url_for('new_models'))

        # Perform predictions based on the selected model
        if selected_model == 'tensorflow_neural':
            model = pickle.load(open('new_models/new-tensorflow_nn.pkl', 'rb'))
            predictions = model.predict(df)
            predictions = (predictions >= 0.5).astype(int)
            func = np.vectorize(churn)
            predictions = func(predictions)
        else:
            if selected_model == 'knn':
                model = pickle.load(open('new_models/new-KNN.pkl', 'rb'))
            elif selected_model == 'logistic':
                model = pickle.load(open('new_models/new-lg.pkl', 'rb'))
            elif selected_model == 'random_forest':
                model = pickle.load(open('new_models/new-random_forest.pkl', 'rb'))
            elif selected_model == 'svm':
                model = pickle.load(open('new_models/new-svm.pkl', 'rb'))
            else:
                return redirect(url_for('predict'))
            predictions = model.predict(df)
            func = np.vectorize(churn)
            predictions = func(predictions)
        if len(predictions) == 0:
            return redirect(url_for('home'))
        
        # Add predictions column to DataFrame
        original_df = pd.read_csv(filepath)
        original_df = original_df['customerID'].to_frame()
        original_df['predictions'] = predictions

        # user file/output prep
        results_filename = "new-" + selected_model + "_results.xlsx"
        results_filepath = os.path.join(app.config['RESULTS_FOLDER'], results_filename)
        original_df['No.'] = range(1, len(df) + 1)
        original_df.set_index('No.', inplace = True)
        original_df.to_excel(results_filepath, index = True)
        session['results_filepath'] = results_filepath
        return render_template('results.html')
    else:
        return render_template('new_models.html')
        
@app.route('/index')
def index():
    return render_template('index.html')

@app.route('/upload_file', methods=['GET', 'POST'])
def upload_file():
    if 'file' not in request.files:
        return redirect(url_for('index'))

    file = request.files['file']
    if file.filename == '':
        return redirect(url_for('index'))

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        session['filepath'] = 'uploads/' + filename
        session['filename'] = filename
        return render_template("pick_models.html")
    else:
        flash("File Error! Submit Only CSV file formats")
        return redirect(url_for('index'))

@app.route('/predict', methods=['GET', 'POST'])
def predict():
    if request.method == 'POST':
        filepath = session.get('filepath')
        if "csv" not in filepath:
            return redirect(url_for('index'))
        
        df = process_files(filepath)

        # Get the selected model option
        selected_model = request.form.get('predict_model')

        if selected_model is None:
            return redirect(url_for('predict'))

        # Perform predictions based on the selected model
        if selected_model == 'tensorflow_neural':
            model = pickle.load(open('models/tensorflow_nn.pkl', 'rb'))
            predictions = model.predict(df)
            predictions = (predictions >= 0.5).astype(int)
            func = np.vectorize(churn)
            predictions = func(predictions)
        else:
            if selected_model == 'knn':
                model = pickle.load(open('models/kNN.pkl', 'rb'))
            elif selected_model == 'logistic':
                model = pickle.load(open('models/lg.pkl', 'rb'))
            elif selected_model == 'random_forest':
                model = pickle.load(open('models/Random_Forest.pkl', 'rb'))
            elif selected_model == 'svm':
                model = pickle.load(open('models/svm_grid.pkl', 'rb'))
            else:
                return redirect(url_for('predict'))
            predictions = model.predict(df)
            func = np.vectorize(churn)
            predictions = func(predictions)
        if len(predictions) == 0:
            return redirect(url_for('home'))

        # Add predictions column to DataFrame
        original_df = pd.read_csv(filepath)
        original_df = original_df['customerID'].to_frame()
        original_df['predictions'] = predictions

        # user file/output prep
        results_filename = selected_model + "_results.xlsx"
        results_filepath = os.path.join(app.config['RESULTS_FOLDER'], results_filename)
        original_df['No.'] = range(1, len(df) + 1)
        original_df.set_index('No.', inplace = True)
        original_df.to_excel(results_filepath, index = True)
        session['results_filepath'] = results_filepath
        return render_template('results.html')
    else:
        return render_template('predict.html')
    
@app.route('/download_results', methods=['GET', 'POST'])
def download_results():
    results_filepath = session.get('results_filepath')
    if request.method == 'GET':
        if not os.path.exists(results_filepath):
            return redirect(url_for('download_results'))
        else:
            return send_file(results_filepath, as_attachment=True)
        
@app.route('/view_results', methods = ['GET', 'POST'])
def view_results():
    filepath = session.get('results_filepath')
    book = load_workbook(filepath)
    sheet = book.active
    if request.method == 'GET':
        return render_template('view_results.html', data = sheet)
    else:
        return redirect(url_for('download_results'))

if __name__ == "__main__":
    app.run(host='0.0.0.0',port=7878, debug = True)